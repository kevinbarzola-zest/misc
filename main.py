import datetime
import openpyxl
import pandas as pd
import time
import xlwings

wb_obj = openpyxl.load_workbook('EquityPrices_ACT.xlsx')
sheet_obj = wb_obj.active
print(sheet_obj)

i = 2
last_cell = sheet_obj.cell(row=i, column=1)
while last_cell.value:
    print(last_cell.value)
    print(i)
    #last_cell.value = datetime.datetime.today()
    i += 1
    last_cell = sheet_obj.cell(row=i, column=1)
wb_obj.save("sample.xlsx")

wb_obj = xlwings.Book("sample.xlsx")
wb_obj.app.calculate()
time.sleep(10)
wb_obj.close()

wb_input = openpyxl.load_workbook('sample.xlsx', data_only=True)
wb_output = openpyxl.load_workbook('sample.xlsx')

ws_input = wb_input.active
ws_output = wb_output.active
for j in range(2, i):
    if type(ws_input.cell(row=i, column=7).value) in [int, float]:
        ws_output.cell(row=i, column=3).value = ws_input.cell(row=i, column=7).value
wb_output.save("sample_1.xlsx")



